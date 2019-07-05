import os
import openpyxl

 

excel='spelling.xlsx'
wb = openpyxl.load_workbook(filename=excel)
ws = wb['Sheet2']
max_row = ws.max_row
path = os.path.dirname(os.path.realpath(__file__))
files = [name for name in os.listdir(path)]
for i in files:
    filepath=os.path.join(path,i)
    for j in range(2,max_row):
        cell_obj=ws.cell(row=j,column=6)
        if cell_obj.value in i:
            os.remove(filepath)
            break